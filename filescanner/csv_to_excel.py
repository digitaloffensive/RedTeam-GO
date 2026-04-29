#!/usr/bin/env python3
"""
csv_to_excel.py  —  Convert FileScanner CSV output to a rich Excel workbook
with embedded screenshot images, colour-coded severity, frozen headers,
auto-filters, and a summary dashboard sheet.

Usage:
    python csv_to_excel.py scan_results.csv
    python csv_to_excel.py scan_results.csv --out report.xlsx
    python csv_to_excel.py scan_results.csv --no-images   # skip image rendering
"""

import argparse
import base64
import csv
import io
import os
import sys
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    print("[warn] Pillow not found — screenshot images will be skipped. "
          "Install with: pip install Pillow")

# ── Colour palette ────────────────────────────────────────────────────────────
SEVERITY_COLORS = {
    "CRITICAL": "C0392B",   # deep red
    "HIGH":     "E67E22",   # orange
    "MEDIUM":   "F1C40F",   # amber
    "LOW":      "27AE60",   # green
}
SEVERITY_TEXT = {
    "CRITICAL": "FFFFFF",
    "HIGH":     "FFFFFF",
    "MEDIUM":   "000000",
    "LOW":      "FFFFFF",
}

HEADER_FILL   = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT   = Font(bold=True, color="E2E8F0", name="Calibri", size=10)
ALT_ROW_FILL  = PatternFill("solid", fgColor="F7F9FC")
NORMAL_FILL   = PatternFill("solid", fgColor="FFFFFF")
BORDER_SIDE   = Side(style="thin", color="D0D7E2")
CELL_BORDER   = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)

# ── Column definitions ────────────────────────────────────────────────────────
COLUMNS = [
    ("ScanDate",          18, "Date/time of detection"),
    ("SharePath",         22, "Root share scanned"),
    ("Folder",            30, "Parent folder"),
    ("FileName",          22, "File name"),
    ("FileExtension",      8, "Ext"),
    ("PatternName",        22, "Pattern that triggered"),
    ("Severity",           10, "Risk level"),
    ("LineNumbers",        12, "Matched lines"),
    ("LinePreview",        40, "First matching line (redacted)"),
    ("Permissions",        26, "File permissions & owner"),
    ("Owner",              14, "File owner"),
    ("FileSizeBytes",      14, "Size in bytes"),
    ("Screenshot",          0, ""),   # image column — width set dynamically
]

IMG_COL_IDX   = 13        # 1-based column index for screenshot
IMG_ROW_HEIGHT = 120      # points
IMG_COL_WIDTH  = 60       # characters (approx)
FONT_SIZE_PX   = 12
FONT_NAME      = "Courier New"


# ── Text → PNG renderer ───────────────────────────────────────────────────────

def text_to_png_bytes(text: str) -> bytes | None:
    """Render mono-spaced text as a PNG image and return raw bytes."""
    if not PIL_AVAILABLE:
        return None

    lines = text.splitlines()
    if not lines:
        return None

    # Approximate dimensions
    char_w, char_h = 7, 14          # pixels per character at ~12pt Courier
    padding = 8
    width  = max(len(l) for l in lines) * char_w + padding * 2
    height = len(lines) * char_h + padding * 2
    width  = max(width, 300)

    img  = PILImage.new("RGB", (width, height), color=(30, 30, 46))
    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", FONT_SIZE_PX)
        font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono-Bold.ttf", FONT_SIZE_PX)
    except Exception:
        font = ImageFont.load_default()
        font_bold = font

    y = padding
    for line in lines:
        # Highlight the ">>>" marker lines in amber
        if line.startswith(">>>"):
            color = (255, 200, 80)
            f = font_bold
        elif line.startswith("===") or line.startswith("| FILE"):
            color = (130, 200, 255)
            f = font_bold
        else:
            color = (200, 210, 230)
            f = font

        draw.text((padding, y), line, fill=color, font=f)
        y += char_h

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


# ── Helpers ───────────────────────────────────────────────────────────────────

def sev_fill(sev: str) -> PatternFill:
    return PatternFill("solid", fgColor=SEVERITY_COLORS.get(sev, "CCCCCC"))

def sev_font(sev: str) -> Font:
    return Font(bold=True, color=SEVERITY_TEXT.get(sev, "000000"),
                name="Calibri", size=10)

def thin_border() -> Border:
    return CELL_BORDER

def style_cell(cell, row_idx: int, bold=False, center=False,
               fill=None, font=None):
    cell.border = thin_border()
    cell.alignment = Alignment(
        wrap_text=True,
        vertical="top",
        horizontal="center" if center else "left"
    )
    if fill:
        cell.fill = fill
    elif row_idx % 2 == 0:
        cell.fill = ALT_ROW_FILL
    else:
        cell.fill = NORMAL_FILL
    if font:
        cell.font = font
    else:
        cell.font = Font(
            bold=bold, name="Calibri", size=10,
            color="1A1A2E"
        )


# ── Findings sheet ────────────────────────────────────────────────────────────

def build_findings_sheet(ws, rows: list[dict], render_images: bool):
    ws.title = "Findings"
    ws.freeze_panes = "A2"

    # Header row
    col_defs = COLUMNS[:-1]  # exclude placeholder screenshot col
    for ci, (col_name, width, _) in enumerate(col_defs, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Screenshot column header
    ss_col = len(col_defs) + 1
    cell = ws.cell(row=1, column=ss_col, value="Screenshot")
    cell.fill  = HEADER_FILL
    cell.font  = HEADER_FONT
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(ss_col)].width = IMG_COL_WIDTH
    ws.row_dimensions[1].height = 20

    ws.auto_filter.ref = f"A1:{get_column_letter(ss_col)}1"

    # Data rows
    for ri, row in enumerate(rows, start=2):
        sev = row.get("Severity", "LOW").upper()

        # Set row height to accommodate image (or compact if no image)
        has_ss = bool(row.get("Screenshot_Base64", "").strip())
        ws.row_dimensions[ri].height = IMG_ROW_HEIGHT if (has_ss and render_images) else 30

        col_keys = [c[0] for c in col_defs]
        for ci, key in enumerate(col_keys, start=1):
            val = row.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)

            if key == "Severity":
                style_cell(cell, ri, bold=True, center=True,
                           fill=sev_fill(sev), font=sev_font(sev))
            else:
                style_cell(cell, ri)

        # Screenshot cell — embed image or leave empty
        ss_cell = ws.cell(row=ri, column=ss_col, value="")
        style_cell(ss_cell, ri)

        b64 = row.get("Screenshot_Base64", "").strip()
        if b64 and render_images and PIL_AVAILABLE:
            try:
                raw_text = base64.b64decode(b64).decode("utf-8", errors="replace")
                png_bytes = text_to_png_bytes(raw_text)
                if png_bytes:
                    img_buf = io.BytesIO(png_bytes)
                    xl_img  = XLImage(img_buf)
                    # Scale to fit row height
                    scale = (IMG_ROW_HEIGHT * 0.95) / xl_img.height if xl_img.height > 0 else 1
                    xl_img.width  = int(xl_img.width  * scale)
                    xl_img.height = int(xl_img.height * scale)
                    cell_ref = f"{get_column_letter(ss_col)}{ri}"
                    ws.add_image(xl_img, cell_ref)
            except Exception as e:
                ss_cell.value = f"[image error: {e}]"
        elif b64 and not PIL_AVAILABLE:
            ss_cell.value = "[install Pillow to see image]"
        elif b64 and not render_images:
            # Store decoded text as cell value instead
            try:
                ss_cell.value = base64.b64decode(b64).decode("utf-8", errors="replace")[:500]
            except Exception:
                pass


# ── Summary dashboard sheet ───────────────────────────────────────────────────

def build_summary_sheet(ws, rows: list[dict]):
    ws.title = "Summary"

    def hdr(cell_ref, text):
        c = ws[cell_ref]
        c.value = text
        c.font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        c.fill  = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    def val(cell_ref, v, bold=False, color="1A1A2E"):
        c = ws[cell_ref]
        c.value = v
        c.font  = Font(bold=bold, name="Calibri", size=11, color=color)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = NORMAL_FILL

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # ── Title block
    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "FILE SHARE SCAN — EXECUTIVE SUMMARY"
    title.font  = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
    title.fill  = PatternFill("solid", fgColor="0D1B2A")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sub.font  = Font(italic=True, color="6B7280", name="Calibri", size=10)
    sub.fill  = PatternFill("solid", fgColor="F0F4F8")
    sub.alignment = Alignment(horizontal="center")

    # ── Severity counts
    sev_counts = defaultdict(int)
    for r in rows:
        sev_counts[r.get("Severity", "LOW").upper()] += 1

    ws.row_dimensions[4].height = 20
    for col, sev in enumerate(["CRITICAL", "HIGH", "MEDIUM", "LOW"], start=1):
        letter = get_column_letter(col)
        hdr(f"{letter}4", sev)
        c = ws[f"{letter}5"]
        c.value = sev_counts.get(sev, 0)
        c.font  = Font(bold=True, name="Calibri", size=22,
                       color=SEVERITY_COLORS.get(sev, "000000"))
        c.fill  = PatternFill("solid", fgColor="F7F9FC")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        ws.row_dimensions[5].height = 36

    # ── Top files by finding count
    file_counts = defaultdict(int)
    for r in rows:
        file_counts[r.get("FileName", "??")] += 1

    ws["A7"].value = "Top Files by Finding Count"
    ws["A7"].font  = Font(bold=True, name="Calibri", size=11, color="1A1A2E")
    ws.merge_cells("A7:C7")

    hdr("A8", "File Name")
    hdr("B8", "Findings")
    hdr("C8", "Share Path")

    top_files = sorted(file_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    for i, (fname, count) in enumerate(top_files, start=9):
        share = next((r["SharePath"] for r in rows if r["FileName"] == fname), "")
        ws[f"A{i}"].value = fname
        ws[f"B{i}"].value = count
        ws[f"C{i}"].value = share
        for col in ["A", "B", "C"]:
            c = ws[f"{col}{i}"]
            c.border = thin_border()
            c.font   = Font(name="Calibri", size=10)
            c.fill   = ALT_ROW_FILL if i % 2 == 0 else NORMAL_FILL
            c.alignment = Alignment(horizontal="left" if col != "B" else "center")

    # ── Pattern breakdown
    row_start = 9 + len(top_files) + 2
    ws[f"A{row_start}"].value = "Pattern Breakdown"
    ws[f"A{row_start}"].font  = Font(bold=True, name="Calibri", size=11)
    ws.merge_cells(f"A{row_start}:C{row_start}")

    hdr(f"A{row_start+1}", "Pattern Name")
    hdr(f"B{row_start+1}", "Count")
    hdr(f"C{row_start+1}", "Severity")

    pattern_counts = defaultdict(lambda: {"count": 0, "sev": "LOW"})
    for r in rows:
        p = r.get("PatternName", "Unknown")
        pattern_counts[p]["count"] += 1
        pattern_counts[p]["sev"]    = r.get("Severity", "LOW")

    for i, (pname, info) in enumerate(
        sorted(pattern_counts.items(), key=lambda x: x[1]["count"], reverse=True),
        start=row_start+2
    ):
        sev = info["sev"].upper()
        ws[f"A{i}"].value = pname
        ws[f"B{i}"].value = info["count"]
        ws[f"C{i}"].value = sev
        for col in ["A", "B", "C"]:
            c = ws[f"{col}{i}"]
            c.border = thin_border()
            c.font   = Font(name="Calibri", size=10)
            c.fill   = sev_fill(sev) if col == "C" else (
                ALT_ROW_FILL if i % 2 == 0 else NORMAL_FILL)
            c.alignment = Alignment(horizontal="center" if col in ["B","C"] else "left")
        ws[f"C{i}"].font = sev_font(sev)


# ── Main ──────────────────────────────────────────────────────────────────────

def convert(csv_path: str, out_path: str, render_images: bool):
    print(f"[*] Reading CSV: {csv_path}")
    rows = []
    # Using utf-8-sig handles BOMs; errors="replace" prevents crashes on bad bytes
    try:
        with open(csv_path, newline="", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
    except Exception as e:
        # Fallback for older Windows-style encodings if utf-8 fails entirely
        print(f"[!] UTF-8 failed, attempting Latin-1 encoding... ({e})")
        rows = []
        with open(csv_path, newline="", encoding="latin-1") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)

    print(f"[*] {len(rows)} findings loaded")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Summary dashboard
    ws_summary = wb.create_sheet("Summary")
    build_summary_sheet(ws_summary, rows)

    # Sheet 2: All findings with screenshots
    ws_findings = wb.create_sheet("Findings")
    build_findings_sheet(ws_findings, rows, render_images)

    wb.save(out_path)
    size_kb = os.path.getsize(out_path) // 1024
    print(f"[✓] Saved: {out_path}  ({size_kb} KB)")


def main():
    parser = argparse.ArgumentParser(
        description="Convert FileScanner CSV to Excel with embedded screenshots"
    )
    parser.add_argument("csv", help="Input CSV file from scanner")
    parser.add_argument("--out", help="Output .xlsx path (default: same name as CSV)")
    parser.add_argument("--no-images", action="store_true",
                        help="Store screenshot text in cell instead of rendering image")
    args = parser.parse_args()

    if not os.path.exists(args.csv):
        print(f"Error: CSV file not found: {args.csv}", file=sys.stderr)
        sys.exit(1)

    out = args.out or os.path.splitext(args.csv)[0] + ".xlsx"
    convert(args.csv, out, render_images=not args.no_images)


if __name__ == "__main__":
    main()
